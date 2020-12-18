#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
@File  : 222.py
@Author: rebecca
@Date  : 2020/9/4 17:15
@Desc  : 
"""
# # 找相同
# list1 = [1, 2, 3, 4]
# list2 = [2, 3, 4, 5]
# print(list1+list2)
# a = [x for x in list1 if x in list2]
# print(a)
# # 找不同
# b = [y for y in (list1 + list2) if y not in a]
# print(b)
#
# # 找出在List1而不在List2中的：
# c = [z for z in list1 if z not in list2]
# print(c)
#
# p = [1,2,3,4]
# p1= [1,2,3,4]
# print(p==p1)
#
# # append方法
# # one_list = [10, "李佳", "Lemon", None]
# # print(id(one_list))  # id是一样的
# # one_var = ["李佳", "Python", "Automated", "Testing"]
# # one_list.append(one_var)
# # print(one_list) # [10, '李佳', 'Lemon', None, ['李佳', 'Python', 'Automated', 'Testing']]
# # print(id(one_list))
#
# # extend方法
# one_list = [10, "李佳", "Lemon", None, (1,2,3,4)]
# one_var = ["Python", "Automated", "Testing"]
# one_list.extend(one_var)
# print(one_list)
# print(id(one_list))
# print(one_list[4])
#
# # 连接符号
# one_list = [10, "李佳", "Lemon", None, (1,2,3,4)]
# one_var = ["李佳", "Python", "Automated", "Testing"]
# print(one_list+one_var)
#
# # 5.删除如下列表中的"矮穷丑"，写出能想到的所有方法
# first_list = [10, '矮矬穷', '李佳', 'Lemon', None, (1, 2, 3, 4), '李佳', 'Python', 'Automated', 'Testing']
# print(first_list)
# del first_list[1]
# print(first_list)
# first_list.pop(1)
# print(first_list)
#
# print((1, 'abc') == (1.0, 'abc'))
# hash((1, 'abc'))
#
# x = (1,2,3)
# l = [1,2,3]
# for index in range(len(x)):
#     print(x[index])
#
# for i in x:
#     print(i)
# print(x.__sizeof__())
# print(l.__sizeof__())
# for row in range(2,  4+1):
#     print(row)

import openpyxl

wb = openpyxl.load_workbook("test222.xlsx")
ws1 = wb.create_sheet("Mysheet444")  #默认最后一个
# print(ws1)

ws1.append([1, 2, 3])
sheet = wb['one'] # 找到sheet页
# print(sheet[1])
# print(sheet)
# print(sheet['A2'].value)

"""
sheet.rows为生成器, 里面是每一行的数据，每一行又由一个tuple包裹。
sheet.columns类似，不过里面是每个tuple是每一列的单元格。
"""
# # 因为按行，所以返回A1, B1, C1这样的顺序
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)
#
# # A1, A2, A3这样的顺序
# for column in sheet.columns:
#     for cell in column:
#         print(cell.value)

rows = sheet.max_row
columns = sheet.max_column
"""
1、获取标题
2、标题就是第一列数据
"""

title = [t.value for t in sheet[1]]

"""
1、获取除标题意外的所有数据
2、标题就是第一列数据
"""
all_data = []
for i in range(2, rows+1):
    row_data = []
    for j in range(1,columns+1):
        row_data.append(sheet.cell(i,j).value)
    # print(row_data)
    row_data = dict(zip(title, row_data))
    """
    1、title为一个标题列表2、数据也为一个列表3、将这个转换为字典格式：dict(zip(list1,list2))
    """
    all_data.append(row_data)
print(all_data)

# print("*******************************************")
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)


all_data = []
for i in range(1, rows+1):
    row_data = []
    for j in range(1,columns+1):
        row_data.append(sheet.cell(i,j).value)
    # print(row_data)
    all_data.append(row_data)
print(all_data)

# 全部是列表格式
list1 = ['key1', 'key2', 'key3']
list2 = ['1', '2', '3']
print(dict(zip(list1, list2)))
wb.save("test222.xlsx")

user_0 = {
    'username': 'efermi',
    'first': 'enrico',
    'last': 'fermi'
}

for key, value in user_0.items():
    print("\nKey:" + key)
    print("Value:" + value)
